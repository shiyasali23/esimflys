#!/usr/bin/env python3
"""
generate_catalog.py — eSIMFlys catalogue cleaner + JSON generator.

Reads the launch catalogue spreadsheet and produces a validated, typed JSON
catalogue for the Next.js frontend to consume at BUILD time (SSG/ISR).

Design decisions (see esim_frontend_design.md):
- KEEP ALL COLUMNS. Nothing is stripped from `plans` — including wholesale_price_usd
  and competitor_ref_*. These are BUILD/SERVER-ONLY fields; the frontend must never
  pass them into a Client Component or render them. The generator preserves them so the
  data stays complete and auditable; safety is enforced at the render layer, not here.
- TWO-TABLE MODEL: `plans` (every row, all columns + derived) and `countries`
  (one row per country: identity + aggregates + editorial placeholders for the SEO gate).
- NO FABRICATION. `status` is preserved as-is (all rows are currently 'paused' — the
  script warns loudly and computes `isLive`, but never flips status). `hotspot` is
  preserved (all 'Unknown' → `hotspotSupported: null`, so the UI hides the claim).
- NORMALIZATION only where safe: country_code → ISO-2 (TUR→TR, etc.), slug, flag emoji,
  numeric casting. Original raw values are retained alongside derived ones.

Usage:
    python scripts/generate_catalog.py \
        --input eSIM_DB_Catalogue_Launch.xlsx \
        --output data/catalog.json

Re-runnable: safe to run on every catalogue update (or as an npm `prebuild` step).
Requires: openpyxl, pycountry  (pip install openpyxl pycountry)
"""
from __future__ import annotations
import argparse, json, re, sys, collections
from pathlib import Path

try:
    import openpyxl
    import pycountry
except ImportError as e:  # pragma: no cover
    sys.exit(f"Missing dependency: {e}. Run: pip install openpyxl pycountry")

# --- Config -----------------------------------------------------------------

# Columns whose VALUES must never reach the browser. Kept in the JSON (per spec),
# but flagged here so the frontend/data layer can project them out for the client.
SERVER_ONLY_FIELDS = ["wholesale_price_usd", "competitor_ref_price", "competitor_ref_brand"]

# Explicit ISO-2 overrides for codes the sheet gets wrong or that pycountry misses.
ISO_OVERRIDES = {"TUR": "TR"}

# Provisional "popular" set (editorial — needs business sign-off, see §38 approvals).
# Only applied to countries actually present in the catalogue; never invents a country.
POPULAR_ISO2 = {"JP", "US", "GB", "FR", "TH", "AE", "TR", "SA", "ID", "IT",
                "ES", "GR", "PT", "MA", "SG", "AU", "DE", "MX"}

NUMERIC_INT = {"day_count", "data_gb", "validity_days", "sort_order"}
NUMERIC_FLOAT = {"wholesale_price_usd", "retail_price_usd", "competitor_ref_price"}


# --- Helpers ----------------------------------------------------------------

def slugify(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")

def to_iso2(raw: str) -> tuple[str, list[str]]:
    """Return (iso2, warnings). Normalizes 3-letter or odd codes to ISO-2."""
    warns: list[str] = []
    code = (raw or "").strip().upper()
    if code in ISO_OVERRIDES:
        return ISO_OVERRIDES[code], warns
    if len(code) == 2:
        if pycountry.countries.get(alpha_2=code) is None:
            warns.append(f"country_code '{code}' is not a known ISO-2 code")
        return code, warns
    # try to resolve 3-letter / numeric / name
    c = (pycountry.countries.get(alpha_3=code)
         or pycountry.countries.get(numeric=code.zfill(3) if code.isdigit() else "000"))
    if c:
        warns.append(f"normalized non-ISO-2 country_code '{raw}' → '{c.alpha_2}'")
        return c.alpha_2, warns
    warns.append(f"could NOT resolve country_code '{raw}' to ISO-2 — left as-is")
    return code, warns

def flag_emoji(iso2: str) -> str:
    if not iso2 or len(iso2) != 2 or not iso2.isalpha():
        return ""
    return "".join(chr(0x1F1E6 + ord(ch) - ord("A")) for ch in iso2.upper())

def num(v, kind):
    if v in (None, ""):
        return None
    try:
        return int(v) if kind is int else round(float(v), 2)
    except (TypeError, ValueError):
        return None

def split_networks(s: str) -> list[str]:
    # split on commas / " and " only — NOT "/", so names like "KDDI/au 5G" stay intact
    return [n.strip() for n in re.split(r",| and ", s or "") if n.strip()]


# --- Core -------------------------------------------------------------------

def load_rows(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    records = [dict(zip(header, r)) for r in rows[1:] if any(c is not None for c in r)]
    return header, records

def build(records: list[dict]):
    warnings: list[str] = []
    plans: list[dict] = []
    countries: dict[str, dict] = {}

    for rec in records:
        iso2, w = to_iso2(rec.get("country_code"))
        warnings.extend(w)
        name = (rec.get("country_name") or "").strip()
        cslug = slugify(name)
        status = (rec.get("status") or "").strip().lower()
        is_live = status == "active"
        plan_type = (rec.get("plan_type") or "").strip().lower()
        is_unlimited = plan_type == "daily"
        hotspot_raw = (rec.get("hotspot") or "").strip().lower()
        hotspot = True if hotspot_raw == "yes" else False if hotspot_raw == "no" else None

        # keep ALL original columns, casting numerics; add derived fields
        plan = dict(rec)
        for k in NUMERIC_INT:
            if k in plan: plan[k] = num(plan[k], int)
        for k in NUMERIC_FLOAT:
            if k in plan: plan[k] = num(plan[k], float)
        plan.update({
            "iso2": iso2,
            "countrySlug": cslug,
            "isLive": is_live,
            "isUnlimited": is_unlimited,
            "perDayGb": plan.get("data_gb") if is_unlimited else None,
            "hotspotSupported": hotspot,
        })
        plans.append(plan)

        c = countries.setdefault(iso2, {
            "slug": cslug, "iso2": iso2, "name": name,
            "region": (rec.get("region") or "").strip(),
            "flagEmoji": flag_emoji(iso2),
            "popular": iso2 in POPULAR_ISO2,
            "planIds": [], "networks": set(), "badges": set(),
            "dataTiersGb": set(), "validityDaysTiers": set(),
            "hasUnlimited": False,
            # editorial placeholders — REQUIRED before the country page may be indexed (§26/§38)
            "content": {"intro": None, "coverageNotes": None, "faqs": [], "approved": False},
        })
        c["planIds"].append(rec.get("product_id"))
        for n in split_networks(rec.get("network")):
            c["networks"].add(n)
        if rec.get("badge"):
            c["badges"].add(str(rec["badge"]).strip())
        if plan.get("data_gb") is not None and not is_unlimited:
            c["dataTiersGb"].add(plan["data_gb"])
        if plan.get("validity_days") is not None:
            c["validityDaysTiers"].add(plan["validity_days"])
        c["hasUnlimited"] = c["hasUnlimited"] or is_unlimited

    # finalize country aggregates
    out_countries = []
    for iso2, c in countries.items():
        c_plans = [p for p in plans if p["iso2"] == iso2]
        live = [p for p in c_plans if p["isLive"]]
        retails = [p["retail_price_usd"] for p in c_plans if p.get("retail_price_usd") is not None]
        c["planCount"] = len(c_plans)
        c["livePlanCount"] = len(live)
        c["priceFrom"] = min(retails) if retails else None       # from ALL plans (none live yet)
        c["networks"] = sorted(c["networks"])
        c["badges"] = sorted(c["badges"])
        c["dataTiersGb"] = sorted(c["dataTiersGb"])
        c["validityDaysTiers"] = sorted(c["validityDaysTiers"])
        c["hotspot"] = None  # all rows 'Unknown' — do not claim
        del c["planIds"]
        out_countries.append(c)
    out_countries.sort(key=lambda c: (not c["popular"], c["name"]))

    # data-quality gates (surface, never silently fix business state)
    live_total = sum(1 for p in plans if p["isLive"])
    if live_total == 0:
        warnings.append("BLOCKER: 0 plans have status='active' — the storefront will show "
                        "NO plans until the business activates them (all rows are 'paused').")
    if all(p["hotspotSupported"] is None for p in plans):
        warnings.append("hotspot is 'Unknown' for every plan — the UI must NOT claim hotspot support.")

    region_country = collections.Counter(c["region"] for c in out_countries)
    region_plan = collections.Counter(p.get("region") for p in plans)
    meta = {
        "source": "eSIM_DB_Catalogue_Launch.xlsx",
        "note": "Generated by scripts/generate_catalog.py. All columns preserved; "
                "SERVER_ONLY fields must be projected out before reaching the client.",
        "serverOnlyFields": SERVER_ONLY_FIELDS,
        "currency": "USD",
        "countryCount": len(out_countries),
        "planCount": len(plans),
        "livePlanCount": live_total,
        "regionsByCountry": dict(sorted(region_country.items(), key=lambda x: -x[1])),
        "regionsByPlan": dict(sorted(region_plan.items(), key=lambda x: -x[1])),
        "dataTiersGb": sorted({p["data_gb"] for p in plans if p.get("data_gb") is not None}),
        "validityDaysTiers": sorted({p["validity_days"] for p in plans if p.get("validity_days") is not None}),
        "planTypes": dict(collections.Counter(p.get("plan_type") for p in plans)),
        "warnings": sorted(set(warnings)),
    }
    return {"meta": meta, "countries": out_countries, "plans": plans}


def main():
    ap = argparse.ArgumentParser(description="Clean + convert the eSIM catalogue xlsx → JSON.")
    ap.add_argument("--input", default="eSIM_DB_Catalogue_Launch.xlsx")
    ap.add_argument("--output", default="data/catalog.json")
    args = ap.parse_args()

    src = Path(args.input)
    if not src.exists():
        sys.exit(f"Input not found: {src.resolve()}")

    header, records = load_rows(src)
    catalog = build(records)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")

    m = catalog["meta"]
    print(f"✓ Wrote {out}  ({out.stat().st_size // 1024} KB)")
    print(f"  columns kept : {len(header)}  ({', '.join(header)})")
    print(f"  countries    : {m['countryCount']}   plans: {m['planCount']}   live: {m['livePlanCount']}")
    print(f"  regions      : {m['regionsByCountry']}")
    print(f"  data tiers   : {m['dataTiersGb']} GB   validity: {m['validityDaysTiers']} d")
    print(f"  plan types   : {m['planTypes']}")
    if m["warnings"]:
        print("  WARNINGS:")
        for w in m["warnings"]:
            print(f"    - {w}")


if __name__ == "__main__":
    main()
