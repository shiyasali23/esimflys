import json
from datetime import date, datetime, timezone as dt_timezone
from decimal import ROUND_HALF_UP, Decimal

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.catalog.models import CatalogPlan, Country, Supplier

# Known source-data corrections (spec §21). Georgia is miscategorised as Africa in the
# workbook; its regional peers here (Azerbaijan, Kazakhstan) are Asia. Fix the source too.
_REGION_CORRECTIONS = {"GE": "Asia"}

COUNTRY_SHEET = "countries"
PLAN_SHEET = "Catalogue"


def _to_minor(amount):
    return int((Decimal(str(amount)) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _gb_to_mb(data_gb):
    return int((Decimal(str(data_gb)) * 1000).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _parse_hotspot(value):
    if value is None:
        return None
    token = str(value).strip().lower()
    if token in {"yes", "true", "y", "supported", "1"}:
        return True
    if token in {"no", "false", "n", "unsupported", "0"}:
        return False
    return None


def _parse_networks(value):
    if not value:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def _is_yes(value):
    return str(value).strip().lower() in {"yes", "true", "1"}


def _clean_badge(value):
    return value if value in {"popular", "value"} else None


def _clean_homepage_badge(value):
    return value if value in {"popular", "best_value"} else None


def _import_status(value):
    token = (value or "").strip().lower()
    return token if token in {"draft", "paused", "retired"} else "paused"


def _verified_at(value):
    if not value:
        return None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        parsed = date.fromisoformat(str(value).strip()[:10])
    return datetime(parsed.year, parsed.month, parsed.day, tzinfo=dt_timezone.utc)


def _sheet_records(workbook, sheet_name):
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    records = []
    for row in rows:
        if all(cell is None for cell in row):
            continue
        records.append(dict(zip(header, row)))
    return records


def _load_workbook(path):
    try:
        import openpyxl
    except ImportError as exc:
        raise CommandError(
            "openpyxl is required to import an .xlsx workbook. Install it with "
            "`pip install openpyxl`."
        ) from exc

    workbook = openpyxl.load_workbook(path, data_only=True)
    sheets = {name.lower(): name for name in workbook.sheetnames}
    for required in (COUNTRY_SHEET, PLAN_SHEET):
        if required.lower() not in sheets:
            raise CommandError(
                f"Workbook is missing the '{required}' sheet (found: {workbook.sheetnames})."
            )

    countries = [
        {
            "iso2": str(row["iso2"]).strip().upper(),
            "name": row["name"],
            "slug": row["slug"],
            "region": (row.get("region") or "").strip(),
            "flag_emoji": row.get("flag_emoji"),
            "timezone": row.get("timezone"),
            "is_popular": bool(row.get("is_popular")),
            "homepage_badge": _clean_homepage_badge(row.get("homepage_badge")),
            "is_active": True if row.get("is_active") is None else bool(row.get("is_active")),
            "sort_order": int(row.get("sort_order") or 0),
        }
        for row in _sheet_records(workbook, sheets[COUNTRY_SHEET.lower()])
    ]
    plans = [
        {
            "product_code": row["product_id"],
            "supplier_package_code": row["supplier_package_code"],
            "plan_type": row["plan_type"],
            "day_count": row.get("day_count"),
            "iso2": str(row["country_code"]).strip().upper(),
            "display_name": row["display_name"],
            "data_gb": row["data_gb"],
            "validity_days": row["validity_days"],
            "traffic_policy": row.get("traffic_policy"),
            "hotspot": row.get("hotspot"),
            "network": row.get("network"),
            "topup_supported": row.get("topup_supported"),
            "retail_price_usd": row["retail_price_usd"],
            "wholesale_price_usd": row.get("wholesale_price_usd"),
            "status": row.get("status"),
            "badge": row.get("badge"),
            "tier": row.get("tier"),
            "default_selected": row.get("default_selected"),
            "sort_order": row.get("sort_order") or 0,
            "verified_date": row.get("wsp_verified_date"),
        }
        for row in _sheet_records(workbook, sheets[PLAN_SHEET.lower()])
    ]
    return countries, plans


def _load_json(path):
    try:
        with open(path) as handle:
            data = json.load(handle)
    except (OSError, json.JSONDecodeError) as exc:
        raise CommandError(f"Could not read catalogue at {path}: {exc}") from exc

    countries = [
        {
            "iso2": str(row["iso2"]).strip().upper(),
            "name": row["name"],
            "slug": row["slug"],
            "region": (row.get("region") or "").strip(),
            "flag_emoji": row.get("flagEmoji"),
            "timezone": None,
            "is_popular": bool(row.get("popular")),
            "homepage_badge": None,
            "is_active": True,
            "sort_order": index,
        }
        for index, row in enumerate(data.get("countries") or [])
    ]
    plans = [
        {
            "product_code": row["product_id"],
            "supplier_package_code": row["supplier_package_code"],
            "plan_type": row["plan_type"],
            "day_count": row.get("day_count"),
            "iso2": str(row["iso2"]).strip().upper(),
            "display_name": row["display_name"],
            "data_gb": row["data_gb"],
            "validity_days": row["validity_days"],
            "traffic_policy": row.get("traffic_policy"),
            "hotspot": row.get("hotspot"),
            "network": row.get("network"),
            "topup_supported": row.get("topup_supported"),
            "retail_price_usd": row["retail_price_usd"],
            "wholesale_price_usd": row.get("wholesale_price_usd"),
            "status": row.get("status"),
            "badge": row.get("badge"),
            "tier": row.get("tier"),
            "default_selected": row.get("default_selected"),
            "sort_order": row.get("sort_order") or 0,
            "verified_date": row.get("wsp_verified_date"),
        }
        for row in data.get("plans") or []
    ]
    return countries, plans


def _default_path():
    workbook = settings.BASE_DIR / "data" / "eSIM_DB_Catalogue_Launch.xlsx"
    if workbook.exists():
        return workbook
    return settings.BASE_DIR.parent / "data" / "catalog.json"


def load_catalogue_source(path=None):
    """Load the catalogue source and return ``(path, kind, countries, plans)``.

    Accepts either the supplier ``.xlsx`` workbook or a generated ``catalog.json`` and
    normalises both into the same record shape. Shared by the import command and the
    reset/validation services so there is a single reader.
    """
    path = path or _default_path()
    if str(path).lower().endswith((".xlsx", ".xlsm")):
        countries, plans = _load_workbook(path)
        return path, "workbook", countries, plans
    countries, plans = _load_json(path)
    return path, "json", countries, plans


class Command(BaseCommand):
    help = (
        "Import the eSIM catalogue from the supplier workbook (.xlsx) or a generated "
        "catalog.json. Upserts countries and plans by stable key, retires missing plans, "
        "and never activates a plan."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--path", default=None, help="Path to the .xlsx workbook or catalog.json."
        )
        parser.add_argument("--supplier-code", default="esim-access")
        parser.add_argument("--supplier-name", default="eSIM Access")
        parser.add_argument(
            "--dry-run", action="store_true", help="Roll back all changes at the end."
        )

    def handle(self, *args, **options):
        path, source, countries, plans = load_catalogue_source(options["path"])

        if not countries or not plans:
            raise CommandError("Catalogue source is missing countries or plans.")

        warnings = []
        with transaction.atomic():
            supplier = self._ensure_supplier(options["supplier_code"], options["supplier_name"])
            country_map, c_created, c_updated = self._upsert_countries(countries, warnings)
            p_created, p_updated, seen = self._upsert_plans(
                plans, supplier, country_map, warnings
            )
            retired = self._retire_missing(supplier, seen)
            if options["dry_run"]:
                transaction.set_rollback(True)

        self._report(
            path, source, c_created, c_updated, p_created, p_updated, retired,
            warnings, options["dry_run"],
        )

    def _ensure_supplier(self, code, name):
        supplier, _ = Supplier.objects.update_or_create(
            code=code, defaults={"name": name, "status": "active"}
        )
        return supplier

    def _upsert_countries(self, records, warnings):
        country_map, created, updated = {}, 0, 0
        for record in records:
            iso2 = record["iso2"]
            raw_region = record["region"]
            region = _REGION_CORRECTIONS.get(iso2, raw_region)
            if region != raw_region:
                warnings.append(
                    f"Corrected region for {iso2}: '{raw_region}' -> '{region}' "
                    "(spec §21; also fix the source workbook)."
                )
            obj, was_created = Country.objects.update_or_create(
                iso2=iso2,
                defaults={
                    "name": record["name"],
                    "slug": record["slug"],
                    "region": region,
                    "flag_emoji": record["flag_emoji"],
                    "timezone": record["timezone"],
                    "is_popular": record["is_popular"],
                    "homepage_badge": record["homepage_badge"],
                    "is_active": record["is_active"],
                    "sort_order": record["sort_order"],
                },
            )
            country_map[iso2] = obj
            created += was_created
            updated += not was_created
        return country_map, created, updated

    def _upsert_plans(self, records, supplier, country_map, warnings):
        created, updated, seen = 0, 0, set()
        for record in records:
            country = country_map.get(record["iso2"])
            if country is None:
                warnings.append(
                    f"Plan {record['product_code']} references unknown country "
                    f"{record['iso2']}; skipped."
                )
                continue
            plan_type = record["plan_type"]
            data_gb = record["data_gb"]
            wholesale = record["wholesale_price_usd"]
            defaults = {
                "supplier": supplier,
                "country": country,
                "supplier_package_code": record["supplier_package_code"],
                "plan_type": plan_type,
                "day_count": record["day_count"],
                "display_name": record["display_name"],
                "data_limit_mb": _gb_to_mb(data_gb) if plan_type == "fixed" else None,
                "daily_high_speed_mb": _gb_to_mb(data_gb) if plan_type == "daily" else None,
                "validity_days": record["validity_days"],
                "traffic_policy": record["traffic_policy"],
                "activation_policy": None,
                "hotspot_supported": _parse_hotspot(record["hotspot"]),
                "network_names": _parse_networks(record["network"]),
                "topup_supported": _is_yes(record["topup_supported"]),
                "retail_amount_minor": _to_minor(record["retail_price_usd"]),
                "wholesale_amount_minor": (
                    _to_minor(wholesale) if wholesale not in (None, "") else None
                ),
                "currency": "USD",
                "status": _import_status(record["status"]),
                "badge": _clean_badge(record["badge"]),
                "tier": record["tier"],
                "is_default_selected": _is_yes(record["default_selected"]),
                "sort_order": record["sort_order"],
                "supplier_verified_at": _verified_at(record["verified_date"]),
            }
            _, was_created = CatalogPlan.objects.update_or_create(
                product_code=record["product_code"], defaults=defaults
            )
            seen.add(record["product_code"])
            created += was_created
            updated += not was_created
        return created, updated, seen

    def _retire_missing(self, supplier, seen):
        return (
            CatalogPlan.objects.filter(supplier=supplier)
            .exclude(product_code__in=seen)
            .exclude(status="retired")
            .update(status="retired")
        )

    def _report(self, path, source, c_created, c_updated, p_created, p_updated, retired, warnings, dry_run):
        write = self.stdout.write
        write("")
        write(
            self.style.MIGRATE_HEADING(
                f"Catalogue import from {source}: {path}"
                + (" (DRY RUN — rolled back)" if dry_run else "")
            )
        )
        write(f"  countries: {c_created} created, {c_updated} updated")
        write(f"  plans:     {p_created} created, {p_updated} updated, {retired} retired")
        write(f"  active plans now: {CatalogPlan.objects.filter(status='active').count()}")
        if warnings:
            write(self.style.WARNING(f"  warnings ({len(dict.fromkeys(warnings))}):"))
            for message in dict.fromkeys(warnings):
                write(self.style.WARNING(f"    - {message}"))
        write(self.style.SUCCESS("  done."))
